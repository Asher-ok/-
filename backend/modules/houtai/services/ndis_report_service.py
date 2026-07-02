"""NDIS 报告生成服务 - 服务使用报告、财务报告"""
from sqlalchemy.orm import Session, joinedload
from decimal import Decimal
from shared.models import Task, Invoice, Customer, Employee
from shared.models import TaskStatus
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from pathlib import Path
from datetime import datetime
from typing import Optional, Tuple, List


def _parse_hours(val: Optional[str]) -> float:
    """解析时长字符串为浮点数"""
    if not val:
        return 0.0
    try:
        return float(val.replace(",", "."))
    except (ValueError, TypeError):
        return 0.0


def get_service_usage_data(
    db: Session,
    customer_id: Optional[str] = None,
    date_start: Optional[datetime] = None,
    date_end: Optional[datetime] = None,
    ndis_only: bool = True
) -> Tuple[List[dict], dict]:
    """获取服务使用报告数据（用于页面展示和 Excel 生成）"""
    query = db.query(Task).join(Task.customer).filter(
        Task.status.in_([TaskStatus.approved, TaskStatus.completed])
    )
    if ndis_only:
        query = query.filter(Customer.customer_type == "NDIS")
    if customer_id:
        query = query.filter(Task.customer_id == customer_id)
    if date_start:
        query = query.filter(Task.service_time >= date_start)
    if date_end:
        query = query.filter(Task.service_time <= date_end)

    tasks = query.options(
        joinedload(Task.customer),
        joinedload(Task.assigned_employee),
        joinedload(Task.invoice_items)
    ).order_by(Task.service_time.desc()).all()

    rows = []
    total_hours = 0.0
    total_cost = Decimal("0")

    for task in tasks:
        customer = task.customer
        employee = task.assigned_employee
        ndis_num = getattr(customer, "ndis_number", None) or ""
        service_time_str = task.service_time.strftime("%Y-%m-%d") if task.service_time else ""
        hours_val = _parse_hours(task.service_duration_hours) or 1.0
        hours_str = task.service_duration_hours or "1"
        staff = employee.name if employee else "-"
        cost = sum(
            (item.amount or Decimal("0")) for item in (task.invoice_items or [])
        )
        total_hours += hours_val
        total_cost += cost

        rows.append({
            "date": service_time_str,
            "participant": customer.name,
            "ndis_number": ndis_num,
            "service_item": task.service_code or task.title or "-",
            "hours": hours_str,
            "staff": staff,
            "cost": float(cost),
            "status": task.status.value if task.status else "-",
        })

    summary = {
        "total_hours": round(total_hours, 2),
        "total_cost": round(float(total_cost), 2),
    }
    return rows, summary


def get_financial_data(
    db: Session,
    customer_id: Optional[str] = None,
    date_start: Optional[datetime] = None,
    date_end: Optional[datetime] = None,
    ndis_only: bool = True
) -> Tuple[List[dict], dict]:
    """获取财务报告数据"""
    query = db.query(Invoice).join(Invoice.customer)
    if ndis_only:
        query = query.filter(Customer.customer_type == "NDIS")
    if customer_id:
        query = query.filter(Invoice.customer_id == customer_id)
    if date_start:
        query = query.filter(Invoice.invoice_date >= date_start)
    if date_end:
        query = query.filter(Invoice.invoice_date <= date_end)

    invoices = query.options(joinedload(Invoice.customer)).order_by(
        Invoice.invoice_date.desc()
    ).all()

    rows = []
    total_amount = Decimal("0")

    for inv in invoices:
        customer = inv.customer
        ndis_num = getattr(customer, "ndis_number", None) or ""
        date_str = inv.invoice_date.strftime("%Y-%m-%d") if inv.invoice_date else ""
        amt = float(inv.total_amount) if inv.total_amount else 0
        total_amount += inv.total_amount or Decimal("0")

        rows.append({
            "invoice_number": inv.invoice_number,
            "date": date_str,
            "participant": customer.name,
            "ndis_number": ndis_num,
            "total_amount": amt,
            "status": inv.status or "draft",
        })

    summary = {"total_amount": round(float(total_amount), 2)}
    return rows, summary


def generate_service_usage_report(
    db: Session,
    output_path: str,
    customer_id: Optional[str] = None,
    date_start: Optional[datetime] = None,
    date_end: Optional[datetime] = None,
    ndis_only: bool = True
) -> str:
    """生成服务使用报告 Excel（含费用列与汇总行）"""
    rows, summary = get_service_usage_data(
        db, customer_id, date_start, date_end, ndis_only
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Service Usage"

    header_font = Font(bold=True)
    ws.append([
        "Date", "Participant", "NDIS Number", "Service Item (Code)", "Hours",
        "Staff", "Cost", "Status"
    ])
    for cell in ws[1]:
        cell.font = header_font

    for r in rows:
        ws.append([
            r["date"], r["participant"], r["ndis_number"], r["service_item"],
            r["hours"], r["staff"], r["cost"], r["status"]
        ])

    if rows:
        ws.append([])
        ws.append([
            "Total",
            "",
            "",
            "",
            summary["total_hours"],
            "",
            summary["total_cost"],
            ""
        ])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

    for col in ws.columns:
        max_length = max(len(str(cell.value) or "") for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

    wb.save(output_path)
    return output_path


def generate_financial_report(
    db: Session,
    output_path: str,
    customer_id: Optional[str] = None,
    date_start: Optional[datetime] = None,
    date_end: Optional[datetime] = None,
    ndis_only: bool = True
) -> str:
    """生成财务报告 Excel（含汇总行）"""
    rows, summary = get_financial_data(
        db, customer_id, date_start, date_end, ndis_only
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Report"

    header_font = Font(bold=True)
    ws.append([
        "Invoice No", "Date", "Participant", "NDIS Number",
        "Total (AUD)", "Status"
    ])
    for cell in ws[1]:
        cell.font = header_font

    for r in rows:
        ws.append([
            r["invoice_number"], r["date"], r["participant"],
            r["ndis_number"], r["total_amount"], r["status"]
        ])

    if rows:
        ws.append([])
        ws.append([
            "Total", "", "", "", summary["total_amount"], ""
        ])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

    for col in ws.columns:
        max_length = max(len(str(cell.value) or "") for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 25)

    wb.save(output_path)
    return output_path
