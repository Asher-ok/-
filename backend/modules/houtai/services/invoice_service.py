from sqlalchemy.orm import Session, joinedload
from sqlalchemy import exists, or_
from shared.models import Invoice, InvoiceItem, InvoiceItemDict, InvoiceAuditLog, InvoiceServiceCode, Task, TaskServiceItem, Customer, TaskStatus as TaskStatusEnum
from core.database import SessionLocal
from core.config import settings
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from typing import List, Optional
import os
from pathlib import Path
import smtplib
import imaplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.header import Header
# Excel生成相关导入（当前使用的格式）
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from PIL import Image as PILImage


def normalize_item_code(code: Optional[str], reference_code: Optional[str] = None) -> Optional[str]:
    if not code:
        return None
    normalized = str(code).strip()
    if normalized.endswith("...") and reference_code:
        base = normalized[:-3].strip()
        base_parts = [p for p in base.split("_") if p]
        ref_parts = [p for p in str(reference_code).strip().split("_") if p]
        if len(base_parts) >= 2 and len(ref_parts) >= len(base_parts):
            return "_".join(base_parts + ref_parts[len(base_parts):])
        return base
    return normalized


def calculate_invoice_line_amounts(unit_price: Decimal, quantity: Decimal, tax_rate: Decimal):
    amount_excl_tax = (unit_price * quantity).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    tax_amount = (amount_excl_tax * tax_rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    amount_incl_tax = (amount_excl_tax + tax_amount).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return amount_excl_tax, tax_amount, amount_incl_tax


def _to_jsonable(value):
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def _invoice_snapshot(invoice: Invoice):
    items = []
    for it in invoice.items or []:
        items.append({
            "id": it.id,
            "line_no": it.line_no,
            "item_id": it.item_id,
            "category_id": it.category_id,
            "item_code": it.item_code or it.service_code,
            "item_name": it.item_name or it.description,
            "specification": it.specification,
            "unit": it.unit,
            "unit_price": _to_jsonable(it.unit_price or it.price),
            "quantity": _to_jsonable(it.quantity),
            "amount_excl_tax": _to_jsonable(it.amount_excl_tax),
            "tax_rate": _to_jsonable(it.tax_rate),
            "tax_amount": _to_jsonable(it.tax_amount),
            "amount_incl_tax": _to_jsonable(it.amount_incl_tax or it.amount),
            "source_task_id": it.source_task_id or it.task_id,
            "remark": it.remark,
            "service_date": _to_jsonable(it.service_date),
            "service_time_start": it.service_time_start,
            "service_time_end": it.service_time_end,
        })

    return {
        "id": invoice.id,
        "invoice_number": invoice.invoice_number,
        "customer_id": invoice.customer_id,
        "invoice_date": _to_jsonable(invoice.invoice_date),
        "status": invoice.status,
        "currency": invoice.currency,
        "buyer_name": invoice.buyer_name,
        "buyer_phone": invoice.buyer_phone,
        "buyer_email": invoice.buyer_email,
        "buyer_address": invoice.buyer_address,
        "total_excl_tax": _to_jsonable(invoice.total_excl_tax),
        "total_tax": _to_jsonable(invoice.total_tax),
        "total_incl_tax": _to_jsonable(invoice.total_incl_tax),
        "total_amount": _to_jsonable(invoice.total_amount),
        "items": items,
    }


def generate_invoice_number(db: Session) -> str:
    """生成发票编号：INV-YYMMDD-序号"""
    today = datetime.now()
    date_str = today.strftime("%y%m%d")
    
    # 查询当天已有的发票数量
    today_start = datetime(today.year, today.month, today.day)
    count = db.query(Invoice).filter(
        Invoice.invoice_number.like(f"INV-{date_str}%"),
        Invoice.created_at >= today_start
    ).count()
    
    sequence = str(count + 1).zfill(2)
    return f"INV-{date_str}{sequence}"


def generate_invoice_pdf(invoice: Invoice, customer: Customer, output_path: str) -> None:
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )
    story = []
    styles = getSampleStyleSheet()

    base_style = ParagraphStyle(
        "Base",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        textColor=colors.HexColor("#333333"),
        alignment=TA_LEFT,
        leading=12,
    )
    title_style = ParagraphStyle(
        "Title",
        parent=base_style,
        fontSize=22,
        textColor=colors.HexColor("#111111"),
        spaceAfter=10,
        alignment=TA_CENTER,
        leading=26,
    )
    small_style = ParagraphStyle(
        "Small",
        parent=base_style,
        fontSize=9,
        leading=11,
    )
    right_style = ParagraphStyle(
        "Right",
        parent=base_style,
        alignment=TA_RIGHT,
    )

    story.append(Paragraph("TAX INVOICE", title_style))
    story.append(Spacer(1, 6 * mm))

    sender_lines = [
        f"<b>{settings.invoice_company_name}</b>",
        f"ABN: {settings.invoice_abn}",
        settings.invoice_address,
        f"Mobile: {settings.invoice_phone}",
        f"Email: {settings.invoice_email}",
    ]
    sender_para = Paragraph("<br/>".join(sender_lines), base_style)

    logo_flowable = None
    logo_path = Path(__file__).parent.parent.parent.parent / "static" / "logo.png"
    if logo_path.exists():
        try:
            logo_flowable = Image(str(logo_path), width=40 * mm, height=18 * mm)
            logo_flowable.hAlign = "RIGHT"
        except Exception:
            logo_flowable = None

    header_table = Table(
        [[sender_para, logo_flowable or Paragraph("", base_style)]],
        colWidths=[115 * mm, 55 * mm],
    )
    header_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(header_table)
    story.append(Spacer(1, 6 * mm))

    invoice_date = invoice.invoice_date.strftime("%d/%m/%Y") if invoice.invoice_date else ""
    invoice_info_rows = [
        [Paragraph("<b>Invoice Number:</b>", base_style), Paragraph(invoice.invoice_number or "", base_style)],
        [Paragraph("<b>Date:</b>", base_style), Paragraph(invoice_date, base_style)],
    ]
    invoice_info_table = Table(invoice_info_rows, colWidths=[35 * mm, 135 * mm])
    invoice_info_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )
    story.append(invoice_info_table)
    story.append(Spacer(1, 6 * mm))

    story.append(Paragraph("<b>Bill to:</b>", base_style))
    recipient_rows = [[Paragraph("<b>Name:</b>", base_style), Paragraph(customer.name or "", base_style)]]
    ndis_number = getattr(customer, "ndis_number", None)
    if ndis_number:
        recipient_rows.append([Paragraph("<b>NDIS Number:</b>", base_style), Paragraph(str(ndis_number), base_style)])

    recipient_table = Table(recipient_rows, colWidths=[35 * mm, 135 * mm])
    recipient_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )
    story.append(recipient_table)
    story.append(Spacer(1, 6 * mm))

    table_data = [
        [
            Paragraph("<b>Description</b>", small_style),
            Paragraph("<b>Code</b>", small_style),
            Paragraph("<b>Price</b>", small_style),
            Paragraph("<b>Qty</b>", small_style),
            Paragraph("<b>Amount (AUD)</b>", small_style),
        ]
    ]

    for item in invoice.items or []:
        desc_lines = [str(item.description or "")]
        if item.service_date:
            date_str = item.service_date.strftime("%d/%m/%Y")
            if item.service_time_start and item.service_time_end:
                desc_lines.append(f"Date/Time: {date_str} {item.service_time_start}-{item.service_time_end}")
            else:
                desc_lines.append(f"Date/Time: {date_str}")

        description = Paragraph("<br/>".join(desc_lines), small_style)
        code = Paragraph(str(item.item_code or item.service_code or ""), small_style)
        unit_price = Decimal(str(item.unit_price if item.unit_price is not None else (item.price or 0)))
        qty_value = Decimal(str(item.quantity or 0))
        amount_value = Decimal(str(item.amount_incl_tax if item.amount_incl_tax is not None else (item.amount or 0)))
        qty_display = f"{qty_value.normalize():f}".rstrip("0").rstrip(".") if qty_value % 1 != 0 else f"{qty_value:.0f}"
        price = Paragraph(f"${unit_price:.2f}", ParagraphStyle("PriceRight", parent=small_style, alignment=TA_RIGHT))
        qty = Paragraph(qty_display, ParagraphStyle("QtyRight", parent=small_style, alignment=TA_RIGHT))
        amount = Paragraph(f"${amount_value:.2f}", ParagraphStyle("AmtRight", parent=small_style, alignment=TA_RIGHT))
        table_data.append([description, code, price, qty, amount])

    total_qty = sum((Decimal(str(item.quantity or 0)) for item in (invoice.items or [])), Decimal("0"))
    total_qty_display = f"{total_qty.normalize():f}".rstrip("0").rstrip(".") if total_qty % 1 != 0 else f"{total_qty:.0f}"
    table_data.append(
        [
            Paragraph("<b>TOTAL</b>", small_style),
            Paragraph("", small_style),
            Paragraph("", small_style),
            Paragraph(f"<b>{total_qty_display}</b>", ParagraphStyle("TotalQty", parent=small_style, alignment=TA_RIGHT)),
            Paragraph(f"<b>${Decimal(str(invoice.total_amount or 0)):.2f}</b>", ParagraphStyle("TotalAmt", parent=small_style, alignment=TA_RIGHT)),
        ]
    )

    service_table = Table(table_data, colWidths=[80 * mm, 25 * mm, 20 * mm, 20 * mm, 25 * mm], repeatRows=1)
    service_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (1, -1), "LEFT"),
                ("ALIGN", (2, 0), (4, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.8, colors.HexColor("#000000")),
                ("LINEABOVE", (0, 0), (-1, 0), 1.2, colors.HexColor("#000000")),
                ("LINEBELOW", (0, 0), (-1, 0), 1.2, colors.HexColor("#000000")),
                ("LINEBELOW", (0, -1), (-1, -1), 1.2, colors.HexColor("#000000")),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("TOPPADDING", (0, 0), (-1, 0), 6),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ]
        )
    )
    story.append(service_table)
    story.append(Spacer(1, 8 * mm))

    payment_rows = [
        [Paragraph("<b>Bank:</b>", base_style), Paragraph(settings.invoice_bank_name, base_style)],
        [Paragraph("<b>Branch Location:</b>", base_style), Paragraph(settings.invoice_bank_branch, base_style)],
        [Paragraph("<b>Account Name:</b>", base_style), Paragraph(settings.invoice_account_name, base_style)],
        [Paragraph("<b>BSB:</b>", base_style), Paragraph(settings.invoice_bsb, base_style)],
        [Paragraph("<b>Account Number:</b>", base_style), Paragraph(settings.invoice_account_number, base_style)],
    ]
    payment_table = Table(payment_rows, colWidths=[45 * mm, 125 * mm])
    payment_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )
    story.append(payment_table)

    status_value = (getattr(invoice, "status", None) or "").strip().lower()

    def _draw_footer(canvas, doc_obj):
        canvas.saveState()
        x = doc_obj.leftMargin
        y1 = 12 * mm
        y2 = 6 * mm

        canvas.setFillColor(colors.HexColor("#000000"))
        canvas.setLineWidth(1.2)

        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawString(x, y1, "PAYMENT ADVICE")
        canvas.line(x, y1 - 2, x + 60 * mm, y1 - 2)

        if status_value == "paid":
            canvas.setFont("Helvetica", 12)
            canvas.drawString(x, y2, "Client Paid")
            canvas.line(x, y2 - 2, x + 45 * mm, y2 - 2)

        canvas.restoreState()

    doc.build(story, onFirstPage=_draw_footer, onLaterPages=_draw_footer)


def generate_invoice_excel(invoice: Invoice, customer: Customer, output_path: str):
    def _column_width_to_pixels(width: float) -> float:
        # Excel列宽到像素的近似换算（与Excel渲染更接近）
        if width is None:
            return 0
        if width <= 1:
            return width * 12
        return width * 7 + 5

    def _points_to_pixels(points: float) -> float:
        if points is None:
            return 0
        return points * 96 / 72
    """生成发票Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # 设置列宽（匹配样例）
    ws.column_dimensions['A'].width = 13.6640625
    ws.column_dimensions['B'].width = 12.6640625
    ws.column_dimensions['C'].width = 7.21875
    ws.column_dimensions['D'].width = 13.0
    ws.column_dimensions['E'].width = 13.0
    ws.column_dimensions['F'].width = 7.0
    ws.column_dimensions['G'].width = 13.0
    ws.column_dimensions['H'].width = 13.0
    ws.column_dimensions['I'].width = 11.21875
    ws.column_dimensions['J'].width = 13.0
    
    # 样式定义（匹配样例）
    company_title_font = Font(name="Dreaming Outloud Pro", size=20, bold=True)
    company_font = Font(name="Dreaming Outloud Pro", size=11)
    header_label_font = Font(name="Dreaming Outloud Pro", size=11)
    payment_font = Font(name="Dreaming Outloud Pro", size=10)
    payment_title_font = Font(name="Dreaming Outloud Pro", size=12)
    participant_font = Font(name="Aptos Narrow", size=11)
    table_header_font = Font(name="Aptos Narrow", size=11, bold=False)
    table_header_font_bold_aptos = Font(name="Aptos Narrow", size=11, bold=True)
    table_header_font_bold_arial = Font(name="Arial", size=9, bold=True)
    table_row_font = Font(name="Arial", size=9)
    total_label_font = Font(name="Arial", size=9, bold=True)
    total_value_font = Font(name="Aptos Narrow", size=11, bold=True)
    
    border_top = Border(top=Side(style='thin', color='000000'))
    border_bottom_light = Border(bottom=Side(style='thin', color='CCCCCC'))
    border_top_bottom = Border(top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    border_bottom = Border(bottom=Side(style='thin', color='000000'))
    
    # 行高（匹配样例）
    ws.row_dimensions[3].height = 25.8
    ws.row_dimensions[4].height = 24.6
    
    # 插入Logo（占用 F-I 列、1-5 行）
    logo_path = Path(__file__).parent.parent.parent.parent / "static" / "logo.png"
    if logo_path.exists():
        try:
            img = OpenpyxlImage(str(logo_path))
            logo_width_px = sum(
                _column_width_to_pixels(ws.column_dimensions[col].width or 0)
                for col in ["F", "G", "H", "I"]
            )
            default_row_height = ws.sheet_format.defaultRowHeight or 15
            logo_height_px = 0
            for row in range(1, 6):
                height_pt = ws.row_dimensions[row].height or default_row_height
                logo_height_px += _points_to_pixels(height_pt)

            img.width = int(logo_width_px)
            img.height = int(logo_height_px)
            img.anchor = OneCellAnchor(
                _from=AnchorMarker(
                    col=5,
                    colOff=0,
                    row=0,
                    rowOff=0
                ),
                ext=XDRPositiveSize2D(img.width * 9525, img.height * 9525)
            )
            ws.add_image(img)
        except Exception as e:
            print(f"无法插入logo: {e}")
    
    # 公司信息（左侧A3-A9）
    ws['A3'] = settings.invoice_company_name
    ws['A3'].font = company_title_font
    ws['A4'] = f"ABN:  {settings.invoice_abn}"
    ws['A4'].font = company_font
    # 地址处理：根据示例，地址分为两行
    address_parts = settings.invoice_address.split(',')
    if len(address_parts) >= 2:
        ws['A6'] = address_parts[0].strip()
        ws['A7'] = address_parts[-1].strip()
    else:
        ws['A6'] = settings.invoice_address
        ws['A7'] = ""
    ws['A8'] = f"M: {settings.invoice_phone}"
    ws['A9'] = settings.invoice_email
    for addr in ['A6', 'A7', 'A8', 'A9']:
        ws[addr].font = company_font
    
    # 发票头部（右上G3-G7）
    ws['G6'] = "Invoice Number:"
    ws['G6'].font = header_label_font
    ws['I6'] = invoice.invoice_number
    ws['I6'].font = header_label_font
    ws['I6'].alignment = Alignment(horizontal='right')
    ws['G7'] = "Date:"
    ws['G7'].font = header_label_font
    ws['I7'] = invoice.invoice_date.strftime('%d/%m/%Y')
    ws['I7'].font = header_label_font
    ws['I7'].alignment = Alignment(horizontal='right')
    
    # 客户信息（A11-B13）
    ws['A11'] = "Bill to:"
    ws['A11'].font = header_label_font
    ws['A12'] = "Participant :"
    ws['B12'] = customer.name
    ws['A12'].font = participant_font
    ws['B12'].font = participant_font
    ws['B12'].alignment = Alignment(horizontal='left', vertical='top')
    ndis_number = getattr(customer, 'ndis_number', None)
    if ndis_number:
        ws['A13'] = "NDIS NUMBER:"
        ws['B13'] = ndis_number
        ws['A13'].font = participant_font
        ws['B13'].font = participant_font
        ws['A13'].alignment = Alignment(horizontal='left', vertical='top')
        ws['B13'].alignment = Alignment(horizontal='left', vertical='top')
    
    # 服务项目表格（从第15行开始）
    # 定义列位置常量（提高可维护性）
    COL_DESCRIPTION = 1  # A列
    COL_CODE = 4  # D列
    COL_PRICE = 7  # G列
    COL_AMOUNT = 8  # H列（数量）
    COL_AMOUNT_AUD = 9  # I列（金额，使用公式）
    COL_TOTAL_LABEL = 6  # F列（总计标签）
    
    row = 15
    ws.merge_cells("B15:C15")
    ws.merge_cells("D15:E15")
    # 表头 - 根据模板格式：A列(Description), D列(Code), G列(price), H列(Amount), I列(Amount AUD)
    ws.cell(row=row, column=COL_DESCRIPTION).value = "Description"
    ws.cell(row=row, column=COL_DESCRIPTION).font = table_header_font
    ws.cell(row=row, column=COL_DESCRIPTION).alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
    
    ws.cell(row=row, column=COL_CODE).value = "Code"
    ws.cell(row=row, column=COL_CODE).font = table_header_font_bold_arial
    ws.cell(row=row, column=COL_CODE).alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
    
    ws.cell(row=row, column=COL_PRICE).value = "price"
    ws.cell(row=row, column=COL_PRICE).font = table_header_font_bold_aptos
    ws.cell(row=row, column=COL_PRICE).alignment = Alignment(horizontal='center')
    
    ws.cell(row=row, column=COL_AMOUNT).value = "Amount"
    ws.cell(row=row, column=COL_AMOUNT).font = table_header_font_bold_arial
    ws.cell(row=row, column=COL_AMOUNT).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.cell(row=row, column=COL_AMOUNT).number_format = "__@"
    
    ws.cell(row=row, column=COL_AMOUNT_AUD).value = "Amount AUD"
    ws.cell(row=row, column=COL_AMOUNT_AUD).font = table_header_font
    ws.cell(row=row, column=COL_AMOUNT_AUD).alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
    
    # 数据行
    row += 1
    first_data_row = row  # 记录第一行数据行号，用于公式计算
    for item in invoice.items:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 30.0
        
        # Description (A列)
        desc_parts = [item.description]
        if item.service_date:
            # 日期格式：移除前导零（如 01 -> 1, 02 -> 2），与模板保持一致
            day = item.service_date.day
            month = item.service_date.month
            year = item.service_date.year
            date_str = f"{day}/{month}/{year}"
            if item.service_time_start and item.service_time_end:
                desc_parts.append(f"Date/Time:{date_str} {item.service_time_start}-{item.service_time_end}")
            else:
                desc_parts.append(f"Date/Time:{date_str}")
        ws.cell(row=row, column=COL_DESCRIPTION).value = " ".join(desc_parts)
        ws.cell(row=row, column=COL_DESCRIPTION).font = table_row_font
        ws.cell(row=row, column=COL_DESCRIPTION).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Code (D列)
        ws.cell(row=row, column=COL_CODE).value = item.item_code or item.service_code or ""
        ws.cell(row=row, column=COL_CODE).font = table_row_font
        ws.cell(row=row, column=COL_CODE).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=row, column=COL_CODE).number_format = "________@"
        
        # Price (G列) - 数值类型，不是字符串
        ws.cell(row=row, column=COL_PRICE).value = float(item.unit_price if item.unit_price is not None else (item.price or 0))
        ws.cell(row=row, column=COL_PRICE).font = table_row_font
        ws.cell(row=row, column=COL_PRICE).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Amount/Quantity (H列) - 数值类型
        ws.cell(row=row, column=COL_AMOUNT).value = float(item.quantity)
        ws.cell(row=row, column=COL_AMOUNT).font = table_row_font
        ws.cell(row=row, column=COL_AMOUNT).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Amount AUD (I列) - 使用公式 =G{row}*H{row}
        formula = f"=G{row}*H{row}"
        ws.cell(row=row, column=COL_AMOUNT_AUD).value = formula
        ws.cell(row=row, column=COL_AMOUNT_AUD).font = table_row_font
        ws.cell(row=row, column=COL_AMOUNT_AUD).alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        ws.cell(row=row, column=COL_AMOUNT_AUD).number_format = "#,##0.00_ "
        
        # 行边框（仅顶部与底部）
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(top=border_top.top, bottom=border_bottom_light.bottom)
        
        row += 1
    
    last_data_row = row - 1  # 记录最后一行数据行号
    
    # 总计行
    ws.row_dimensions[row].height = 30.0
    # F列: "TOTAL"
    ws.cell(row=row, column=COL_TOTAL_LABEL).value = "TOTAL"
    ws.cell(row=row, column=COL_TOTAL_LABEL).font = total_label_font
    ws.cell(row=row, column=COL_TOTAL_LABEL).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # H列: 总数量公式 =SUM(H{first_row}:H{last_row})
    total_qty_formula = f"=SUM(H{first_data_row}:H{last_data_row})"
    ws.cell(row=row, column=COL_AMOUNT).value = total_qty_formula
    ws.cell(row=row, column=COL_AMOUNT).font = total_value_font
    ws.cell(row=row, column=COL_AMOUNT).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.cell(row=row, column=COL_AMOUNT).number_format = "#,##0_ "
    
    # I列: 总金额公式 =SUM(I{first_row}:I{last_row})
    total_amount_formula = f"=SUM(I{first_data_row}:I{last_data_row})"
    ws.cell(row=row, column=COL_AMOUNT_AUD).value = total_amount_formula
    ws.cell(row=row, column=COL_AMOUNT_AUD).font = total_value_font
    ws.cell(row=row, column=COL_AMOUNT_AUD).alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    ws.cell(row=row, column=COL_AMOUNT_AUD).number_format = "#,##0.00_ "
    
    # 合计行边框（与样例一致）
    for col in range(1, 10):
        cell = ws.cell(row=row, column=col)
        if col in (1, 2, 8, 9):
            cell.border = Border(bottom=border_bottom.bottom)
        else:
            cell.border = border_top_bottom
    
    # 支付信息（紧随合计行）
    blank_row = row + 1
    payment_start_row = row + 2
    ws.merge_cells(start_row=blank_row, start_column=1, end_row=blank_row, end_column=2)
    ws.merge_cells(start_row=payment_start_row, start_column=1, end_row=payment_start_row, end_column=9)
    ws.merge_cells(start_row=payment_start_row + 1, start_column=1, end_row=payment_start_row + 1, end_column=9)
    ws.merge_cells(start_row=payment_start_row + 2, start_column=2, end_row=payment_start_row + 2, end_column=8)
    ws.merge_cells(start_row=payment_start_row + 3, start_column=1, end_row=payment_start_row + 3, end_column=9)
    ws.merge_cells(start_row=payment_start_row + 4, start_column=1, end_row=payment_start_row + 4, end_column=9)
    
    ws[f'A{payment_start_row}'] = settings.invoice_bank_name
    ws[f'A{payment_start_row}'].font = payment_font
    ws[f'A{payment_start_row}'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws[f'A{payment_start_row + 1}'] = f"branch location: {settings.invoice_bank_branch}"
    ws[f'A{payment_start_row + 1}'].font = payment_font
    ws[f'A{payment_start_row + 1}'].alignment = Alignment(horizontal='left', vertical='top')
    ws[f'A{payment_start_row + 1}'].number_format = "@"
    
    ws[f'A{payment_start_row + 2}'] = "Account name:"
    ws[f'A{payment_start_row + 2}'].font = payment_font
    ws[f'A{payment_start_row + 2}'].alignment = Alignment(horizontal='left', vertical='top')
    ws[f'A{payment_start_row + 2}'].number_format = "@"
    ws[f'B{payment_start_row + 2}'] = settings.invoice_account_name
    ws[f'B{payment_start_row + 2}'].font = payment_font
    ws[f'B{payment_start_row + 2}'].alignment = Alignment(horizontal='left', vertical='top')
    ws[f'B{payment_start_row + 2}'].number_format = "@"
    
    ws[f'A{payment_start_row + 3}'] = f"BSB: {settings.invoice_bsb}"
    ws[f'A{payment_start_row + 3}'].font = payment_font
    ws[f'A{payment_start_row + 3}'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws[f'A{payment_start_row + 4}'] = f"Account: {settings.invoice_account_number}"
    ws[f'A{payment_start_row + 4}'].font = payment_font
    ws[f'A{payment_start_row + 4}'].alignment = Alignment(horizontal='left')
    
    # 空行高度与付款状态区（匹配样例）
    ws.row_dimensions[payment_start_row + 4].height = 15.0
    ws.row_dimensions[payment_start_row + 5].height = 15.0
    
    payment_advice_row = payment_start_row + 6
    ws.merge_cells(start_row=payment_advice_row, start_column=3, end_row=payment_advice_row, end_column=5)
    ws.merge_cells(start_row=payment_advice_row, start_column=8, end_row=payment_advice_row, end_column=9)
    ws.merge_cells(start_row=payment_advice_row + 1, start_column=3, end_row=payment_advice_row + 1, end_column=6)
    ws.merge_cells(start_row=payment_advice_row + 1, start_column=8, end_row=payment_advice_row + 1, end_column=9)
    
    ws[f'A{payment_advice_row}'] = "PAYMENT ADVICE"
    ws[f'A{payment_advice_row}'].font = payment_title_font
    ws[f'A{payment_advice_row}'].alignment = Alignment(horizontal='left', vertical='center')
    if invoice.status == "paid":
        ws[f'A{payment_advice_row + 1}'] = "Client Paid"
        ws[f'A{payment_advice_row + 1}'].font = payment_title_font
        ws[f'A{payment_advice_row + 1}'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws.row_dimensions[payment_advice_row].height = 33.6
    ws.row_dimensions[payment_advice_row + 1].height = 33.6
    
    wb.save(output_path)


def create_invoice_from_tasks(
    db: Session,
    customer_id: str,
    task_ids: Optional[List[str]] = None,
    employee_id: Optional[str] = None,
    date_start: Optional[datetime] = None,
    date_end: Optional[datetime] = None,
    service_code: Optional[str] = None,
    task_overrides: Optional[List] = None,
    invoice_date: Optional[datetime] = None,
    is_paid: bool = False,
    commit: bool = True,
) -> Invoice:
    """从任务创建发票（支持筛选条件）"""
    if invoice_date is None:
        invoice_date = datetime.utcnow()

    if date_end and date_end.hour == 0 and date_end.minute == 0 and date_end.second == 0 and date_end.microsecond == 0:
        date_end = date_end + timedelta(days=1) - timedelta(microseconds=1)
    
    # 验证客户
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise ValueError("客户不存在")
    
    # 构建查询
    query = db.query(Task).options(joinedload(Task.service_items)).filter(
        Task.customer_id == customer_id,
        Task.status == TaskStatusEnum.approved
    )
    invoiced_exists = db.query(InvoiceItem.id).filter(
        or_(
            InvoiceItem.task_id == Task.id,
            InvoiceItem.source_task_id == Task.id,
        )
    ).exists()
    query = query.filter(~invoiced_exists)
    
    # 如果提供了task_ids，使用task_ids筛选
    if task_ids:
        invoiced_ids = set()
        for row in db.query(InvoiceItem.task_id).filter(InvoiceItem.task_id.in_(task_ids)).distinct().all():
            if row and row[0]:
                invoiced_ids.add(row[0])
        for row in db.query(InvoiceItem.source_task_id).filter(InvoiceItem.source_task_id.in_(task_ids)).distinct().all():
            if row and row[0]:
                invoiced_ids.add(row[0])
        if invoiced_ids:
            raise ValueError("所选任务已开发票，不能重复开票")
        query = query.filter(Task.id.in_(task_ids))
    else:
        # 否则使用筛选条件
        if employee_id:
            query = query.filter(Task.assigned_employee_id == employee_id)
        if date_start:
            query = query.filter(Task.service_time >= date_start)
        if date_end:
            query = query.filter(Task.service_time <= date_end)
        if service_code:
            query = query.filter(Task.service_code == service_code)
    
    tasks = query.all()
    
    if not tasks:
        raise ValueError("没有找到符合条件的审核通过的任务")

    task_id_list = [t.id for t in tasks if t and getattr(t, "id", None)]
    service_items_by_task_id = {}
    if task_id_list:
        rows = (
            db.query(TaskServiceItem)
            .filter(TaskServiceItem.task_id.in_(task_id_list))
            .order_by(TaskServiceItem.task_id.asc(), TaskServiceItem.created_at.asc())
            .all()
        )
        for row in rows:
            service_items_by_task_id.setdefault(row.task_id, []).append(row)

    override_map = {}
    if task_overrides:
        override_map = {item.task_id: item for item in task_overrides if item and getattr(item, "task_id", None)}

    # 生成发票编号
    invoice_number = generate_invoice_number(db)
    
    # 创建发票
    buyer_name = getattr(customer, "invoice_receiver_name", None) or customer.name
    buyer_phone = getattr(customer, "invoice_receiver_phone", None)
    buyer_email = getattr(customer, "invoice_receiver_email", None) or customer.email
    buyer_address = getattr(customer, "invoice_receiver_address", None) or getattr(customer, "address", None)

    invoice = Invoice(
        invoice_number=invoice_number,
        customer_id=customer_id,
        invoice_date=invoice_date,
        total_amount=Decimal("0"),
        total_excl_tax=Decimal("0"),
        total_tax=Decimal("0"),
        total_incl_tax=Decimal("0"),
        currency="AUD",
        buyer_name=buyer_name,
        buyer_phone=buyer_phone,
        buyer_email=buyer_email,
        buyer_address=buyer_address,
        status="paid" if is_paid else "draft",
        paid_at=invoice_date if is_paid else None,
    )
    db.add(invoice)
    db.flush()
    
    # 创建发票项目
    total_excl_tax = Decimal("0")
    total_tax = Decimal("0")
    total_incl_tax = Decimal("0")
    line_no = 1
    for task in tasks:
        service_items = list(service_items_by_task_id.get(task.id, []) or [])
        if service_items:
            for svc in service_items:
                try:
                    price = Decimal(str(getattr(svc, "unit_price", None) or "0"))
                    quantity = Decimal(str(getattr(svc, "quantity", None) or "0"))
                except Exception:
                    raise ValueError("项目价格或数量无效")

                unit_price = price.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                quantity = quantity.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

                raw_code = getattr(svc, "service_code", None) or task.service_code or "04_104_0125_6_1"
                code_norm = normalize_item_code(raw_code)

                svc_code = (
                    db.query(InvoiceServiceCode)
                    .options(joinedload(InvoiceServiceCode.level3))
                    .filter(InvoiceServiceCode.is_active == True, InvoiceServiceCode.code == code_norm)
                    .first()
                )

                dict_item = None
                if code_norm:
                    dict_item = db.query(InvoiceItemDict).filter(InvoiceItemDict.item_code == code_norm).first()

                tax_rate = Decimal(str(getattr(dict_item, "tax_rate_default", 0) or 0))
                if tax_rate < 0 or tax_rate > 1:
                    tax_rate = Decimal("0")

                amount_override = getattr(svc, "amount", None)
                amount_override_dec = None
                try:
                    if amount_override is not None and str(amount_override).strip() != "":
                        amount_override_dec = Decimal(str(amount_override)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                except Exception:
                    amount_override_dec = None

                if amount_override_dec is not None:
                    amount_incl_tax = amount_override_dec
                    if tax_rate and tax_rate > 0:
                        base = (amount_incl_tax / (Decimal("1") + tax_rate)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                        amount_excl_tax = base
                        tax_amount = (amount_incl_tax - amount_excl_tax).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    else:
                        amount_excl_tax = amount_incl_tax
                        tax_amount = Decimal("0.00")
                else:
                    amount_excl_tax, tax_amount, amount_incl_tax = calculate_invoice_line_amounts(unit_price, quantity, tax_rate)

                service_date = task.service_time
                service_time_start = getattr(svc, "service_time_start", None)
                service_time_end = getattr(svc, "service_time_end", None)

                def _extract_service_date_and_hhmm(value):
                    if value is None:
                        return None, None
                    s = str(value).strip()
                    if not s:
                        return None, None
                    if "T" in s or "-" in s:
                        try:
                            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
                            hhmm = dt.strftime("%H%M")
                            return dt, hhmm
                        except Exception:
                            pass
                        try:
                            dt = datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S")
                            hhmm = dt.strftime("%H%M")
                            return dt, hhmm
                        except Exception:
                            pass
                    if ":" in s:
                        parts = s.split(":")
                        if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
                            return None, f"{int(parts[0]):02d}{int(parts[1]):02d}"
                    if len(s) == 4 and s.isdigit():
                        return None, s
                    return None, None

                start_dt, start_hhmm = _extract_service_date_and_hhmm(service_time_start)
                end_dt, end_hhmm = _extract_service_date_and_hhmm(service_time_end)
                if start_dt:
                    service_date = start_dt
                if start_hhmm:
                    service_time_start = start_hhmm
                if end_hhmm:
                    service_time_end = end_hhmm

                if not service_time_start or not service_time_end:
                    if task.service_start_time and task.service_end_time:
                        service_time_start = task.service_start_time.strftime("%H%M")
                        service_time_end = task.service_end_time.strftime("%H%M")
                    elif task.service_time:
                        service_time_start = task.service_time.strftime("%H%M")

                service_name = getattr(getattr(svc_code, "level3", None), "name", None)
                description = service_name or getattr(dict_item, "item_name", None) or getattr(svc, "remark", None) or getattr(task, "title", None) or "Service"
                unit = getattr(svc, "unit", None) or getattr(svc_code, "unit", None) or getattr(dict_item, "unit_default", None) or "Hour"
                item_name = service_name or getattr(dict_item, "item_name", None) or description or "Service"
                category_id = getattr(dict_item, "category_id", None)
                item_id = getattr(dict_item, "id", None)

                item = InvoiceItem(
                    invoice_id=invoice.id,
                    task_id=task.id,
                    task_service_item_id=getattr(svc, "id", None),
                    line_no=line_no,
                    item_id=item_id,
                    category_id=category_id,
                    item_code=code_norm,
                    item_name=item_name,
                    specification=getattr(dict_item, "spec_default", None),
                    unit=unit,
                    unit_price=unit_price,
                    amount_excl_tax=amount_excl_tax,
                    tax_rate=tax_rate,
                    tax_amount=tax_amount,
                    amount_incl_tax=amount_incl_tax,
                    source_task_id=task.id,
                    description=description,
                    service_code=code_norm,
                    price=unit_price,
                    quantity=quantity,
                    amount=amount_incl_tax,
                    service_date=service_date,
                    service_time_start=service_time_start,
                    service_time_end=service_time_end,
                    remark=getattr(svc, "remark", None),
                )
                db.add(item)
                total_excl_tax += amount_excl_tax
                total_tax += tax_amount
                total_incl_tax += amount_incl_tax
                line_no += 1
        else:
            plans = task.service_plans if isinstance(getattr(task, "service_plans", None), list) else None
            if plans:
                for plan in plans:
                    try:
                        price = Decimal(str(plan.get("unit_price") or "0"))
                        quantity = Decimal(str(plan.get("quantity") or "0"))
                    except Exception:
                        raise ValueError("项目价格或数量无效")

                    unit_price = price.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    quantity = quantity.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

                    raw_code = plan.get("service_code") or task.service_code or "04_104_0125_6_1"
                    code_norm = normalize_item_code(raw_code)

                    dict_item = None
                    if code_norm:
                        dict_item = db.query(InvoiceItemDict).filter(InvoiceItemDict.item_code == code_norm).first()

                    tax_rate = Decimal(str(getattr(dict_item, "tax_rate_default", 0) or 0))
                    if tax_rate < 0 or tax_rate > 1:
                        tax_rate = Decimal("0")

                    amount_override_dec = None
                    try:
                        if plan.get("amount") is not None and str(plan.get("amount")).strip() != "":
                            amount_override_dec = Decimal(str(plan.get("amount"))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    except Exception:
                        amount_override_dec = None

                    if amount_override_dec is not None and tax_rate == 0:
                        amount_excl_tax = amount_override_dec
                        tax_amount = Decimal("0.00")
                        amount_incl_tax = amount_override_dec
                    else:
                        amount_excl_tax, tax_amount, amount_incl_tax = calculate_invoice_line_amounts(unit_price, quantity, tax_rate)

                    service_time_start = plan.get("service_time_start")
                    service_time_end = plan.get("service_time_end")
                    if not service_time_start or not service_time_end:
                        if task.service_start_time and task.service_end_time:
                            service_time_start = task.service_start_time.strftime("%H%M")
                            service_time_end = task.service_end_time.strftime("%H%M")
                        elif task.service_time:
                            service_time_start = task.service_time.strftime("%H%M")

                    description = plan.get("remark") or getattr(dict_item, "item_name", None) or getattr(task, "title", None) or "Service"
                    unit = plan.get("unit") or getattr(dict_item, "unit_default", None) or "Hour"
                    item_name = getattr(dict_item, "item_name", None) or description or "Service"
                    category_id = getattr(dict_item, "category_id", None)
                    item_id = getattr(dict_item, "id", None)

                    item = InvoiceItem(
                        invoice_id=invoice.id,
                        task_id=task.id,
                        line_no=line_no,
                        item_id=item_id,
                        category_id=category_id,
                        item_code=code_norm,
                        item_name=item_name,
                        specification=getattr(dict_item, "spec_default", None),
                        unit=unit,
                        unit_price=unit_price,
                        amount_excl_tax=amount_excl_tax,
                        tax_rate=tax_rate,
                        tax_amount=tax_amount,
                        amount_incl_tax=amount_incl_tax,
                        source_task_id=task.id,
                        description=description,
                        service_code=code_norm,
                        price=unit_price,
                        quantity=quantity,
                        amount=amount_incl_tax,
                        service_date=task.service_time,
                        service_time_start=service_time_start,
                        service_time_end=service_time_end,
                    )
                    db.add(item)
                    total_excl_tax += amount_excl_tax
                    total_tax += tax_amount
                    total_incl_tax += amount_incl_tax
                    line_no += 1
            else:
                override = override_map.get(task.id)
                if override:
                    try:
                        quantity = Decimal(override.quantity)
                    except Exception:
                        raise ValueError("项目数量格式不正确")
                    try:
                        price = Decimal(override.price)
                    except Exception:
                        raise ValueError("项目价格格式不正确")
                    if quantity <= 0 or price < 0:
                        raise ValueError("项目价格或数量无效")
                else:
                    if task.service_duration_hours:
                        try:
                            quantity = Decimal(task.service_duration_hours)
                        except Exception:
                            quantity = Decimal("1")
                    else:
                        quantity = Decimal("1")
                    if task.unit_price is not None and Decimal(str(task.unit_price)) > 0:
                        price = Decimal(str(task.unit_price))
                    else:
                        price = None
                raw_code = task.service_code or "04_104_0125_6_1"
                item_code = normalize_item_code(raw_code)
                dict_item = None
                if item_code:
                    dict_item = db.query(InvoiceItemDict).filter(InvoiceItemDict.item_code == item_code).first()
                tax_rate = Decimal(str(getattr(dict_item, "tax_rate_default", 0) or 0))
                unit = getattr(dict_item, "unit_default", None) or "Hour"
                item_name = getattr(dict_item, "item_name", None) or getattr(task, "title", None) or "Service"
                category_id = getattr(dict_item, "category_id", None)
                item_id = getattr(dict_item, "id", None)
                if price is None:
                    dict_price = getattr(dict_item, "price_default", None)
                    if dict_price is not None and Decimal(str(dict_price)) >= 0:
                        price = Decimal(str(dict_price))
                    else:
                        price = Decimal("60")
                unit_price = price.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                quantity = quantity.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
                amount_excl_tax, tax_amount, amount_incl_tax = calculate_invoice_line_amounts(unit_price, quantity, tax_rate)
                service_time_start = None
                service_time_end = None
                if task.service_start_time and task.service_end_time:
                    service_time_start = task.service_start_time.strftime("%H%M")
                    service_time_end = task.service_end_time.strftime("%H%M")
                elif task.service_time:
                    service_time_start = task.service_time.strftime("%H%M")
                    if task.service_duration_hours:
                        try:
                            duration = float(task.service_duration_hours)
                            from dateutil.relativedelta import relativedelta
                            end_time = task.service_time + relativedelta(hours=duration)
                            service_time_end = end_time.strftime("%H%M")
                        except:
                            pass
                item = InvoiceItem(
                    invoice_id=invoice.id,
                    task_id=task.id,
                    line_no=line_no,
                    item_id=item_id,
                    category_id=category_id,
                    item_code=item_code,
                    item_name=item_name,
                    specification=getattr(dict_item, "spec_default", None),
                    unit=unit,
                    unit_price=unit_price,
                    amount_excl_tax=amount_excl_tax,
                    tax_rate=tax_rate,
                    tax_amount=tax_amount,
                    amount_incl_tax=amount_incl_tax,
                    source_task_id=task.id,
                    description=item_name,
                    service_code=item_code,
                    price=unit_price,
                    quantity=quantity,
                    amount=amount_incl_tax,
                    service_date=task.service_time,
                    service_time_start=service_time_start,
                    service_time_end=service_time_end
                )
                db.add(item)
                total_excl_tax += amount_excl_tax
                total_tax += tax_amount
                total_incl_tax += amount_incl_tax
                line_no += 1
    
    invoice.total_excl_tax = total_excl_tax.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    invoice.total_tax = total_tax.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    invoice.total_incl_tax = total_incl_tax.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    invoice.total_amount = invoice.total_incl_tax

    import json
    audit = InvoiceAuditLog(
        invoice_id=invoice.id,
        action="create_from_tasks",
        actor_type="system",
        before_json=None,
        after_json=json.dumps(_invoice_snapshot(invoice), ensure_ascii=False, default=_to_jsonable),
    )
    db.add(audit)

    if commit:
        db.commit()
    else:
        db.flush()
    db.refresh(invoice)
    return invoice


def send_invoice_email(
    invoice: Invoice,
    customer: Customer,
    file_path: str,
    to_email: Optional[str] = None,
    language: str = "en",
) -> bool:
    """发送发票邮件（支持Excel和PDF格式）"""
    if not settings.smtp_host or not settings.smtp_user:
        raise ValueError("SMTP服务器未配置")
    
    recipient_email = (
        (to_email or "").strip()
        or (getattr(invoice, "buyer_email", None) or "").strip()
        or (getattr(customer, "invoice_receiver_email", None) or "").strip()
        or (getattr(customer, "email", None) or "").strip()
    )
    if not recipient_email:
        raise ValueError("客户邮箱不存在，无法发送邮件")
    
    try:
        # 创建邮件
        msg = MIMEMultipart()
        # 确保From地址与登录用户一致，这样邮件服务器才会保存到已发送文件夹
        from_email = settings.smtp_from_email or settings.smtp_user
        msg['From'] = from_email
        msg['To'] = recipient_email
        normalized_lang = (language or "en").strip().lower()
        if normalized_lang not in ("zh", "en"):
            normalized_lang = "en"

        subject = f"Invoice {invoice.invoice_number} - {settings.invoice_company_name}"
        if normalized_lang == "zh":
            subject = f"发票 {invoice.invoice_number} - {settings.invoice_company_name}"
        msg['Subject'] = str(Header(subject, 'utf-8'))
        msg['Reply-To'] = from_email
        
        # 邮件正文
        total_amount = getattr(invoice, "total_amount", None)
        try:
            amount_str = f"{float(total_amount):.2f}" if total_amount is not None else ""
        except Exception:
            amount_str = str(total_amount) if total_amount is not None else ""

        if normalized_lang == "zh":
            body = (
                f"{customer.name}，您好：\n\n"
                f"请查收附件发票（发票号：{invoice.invoice_number}）。\n\n"
                f"总金额：${amount_str}\n\n"
                f"谢谢。\n\n"
                f"{settings.invoice_company_name}\n"
            )
        else:
            body = (
                f"Dear {customer.name},\n\n"
                f"Please find attached invoice {invoice.invoice_number} for services provided.\n\n"
                f"Total Amount: ${amount_str}\n\n"
                f"Thank you for your business.\n\n"
                f"Best regards,\n"
                f"{settings.invoice_company_name}\n"
            )
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        # 附件 - 支持Excel和PDF格式
        file_ext = Path(file_path).suffix.lower()
        if file_ext == '.xlsx':
            mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename = f'invoice_{invoice.invoice_number}.xlsx'
        else:
            mime_type = 'application/pdf'
            filename = f'invoice_{invoice.invoice_number}.pdf'
        
        with open(file_path, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename={filename}'
            )
            msg.attach(part)
        
        # 发送邮件 - 使用sendmail而不是send_message，确保From地址正确
        server = smtplib.SMTP(settings.smtp_host, settings.smtp_port)
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(settings.smtp_user, settings.smtp_password)
        # 使用sendmail方法，明确指定发件人地址
        # 只发送给客户
        server.sendmail(from_email, [recipient_email], msg.as_string())
        server.quit()
        
        # 通过IMAP保存邮件到已发送文件夹
        try:
            import logging
            logger = logging.getLogger(__name__)
            
            # 推断IMAP服务器地址（通常与SMTP相同或使用mail前缀）
            imap_host = settings.smtp_host.replace('smtp.', 'imap.', 1) if 'smtp.' in settings.smtp_host else settings.smtp_host
            if not imap_host.startswith('imap.'):
                # 如果替换失败，尝试使用mail前缀
                imap_host = settings.smtp_host.replace('smtp.', 'mail.', 1) if 'smtp.' in settings.smtp_host else settings.smtp_host
            
            # 连接到IMAP服务器（尝试SSL端口993）
            try:
                imap = imaplib.IMAP4_SSL(imap_host, 993)
            except:
                # 如果SSL失败，尝试非SSL端口143
                try:
                    imap = imaplib.IMAP4(imap_host, 143)
                    imap.starttls()
                except Exception as e:
                    logger.warning(f"无法连接到IMAP服务器 {imap_host}: {str(e)}")
                    raise
            
            imap.login(settings.smtp_user, settings.smtp_password)
            
            # 查找已发送文件夹
            status, folders = imap.list()
            sent_folder = None
            if status == 'OK':
                # 尝试常见的已发送文件夹名称
                possible_names = ['Sent', 'Sent Mail', 'INBOX.Sent']
                for folder_name in possible_names:
                    try:
                        imap.select(folder_name)
                        sent_folder = folder_name
                        logger.info(f"找到已发送文件夹: {sent_folder}")
                        break
                    except:
                        continue
                
                # 如果标准名称都不存在，遍历所有文件夹查找包含"sent"的
                if not sent_folder:
                    for folder in folders:
                        folder_str = folder.decode('utf-8', errors='ignore')
                        if 'sent' in folder_str.lower():
                            try:
                                # IMAP LIST返回格式通常是: (\\HasNoChildren) "/" "Sent"
                                parts = folder_str.split('"')
                                if len(parts) >= 2:
                                    folder_name = parts[-2]
                                    imap.select(folder_name)
                                    sent_folder = folder_name
                                    logger.info(f"找到已发送文件夹: {sent_folder}")
                                    break
                            except:
                                continue
            
            # 如果还是没找到，尝试直接使用'Sent'
            if not sent_folder:
                try:
                    imap.select('Sent')
                    sent_folder = 'Sent'
                    logger.info(f"使用默认已发送文件夹: {sent_folder}")
                except:
                    pass
            
            if sent_folder:
                # 将邮件保存到已发送文件夹
                msg_bytes = msg.as_bytes()
                logger.info(f"准备保存邮件到文件夹: {sent_folder}, 消息大小: {len(msg_bytes)} 字节")
                # APPEND命令: append(mailbox, flags, date_time, message)
                # flags使用空列表[]，date_time使用None
                result = imap.append(sent_folder, [], None, msg_bytes)
                if result[0] == 'OK':
                    logger.info(f"成功保存邮件到已发送文件夹: {sent_folder}")
                else:
                    logger.warning(f"保存到已发送文件夹返回: {result}")
            else:
                logger.warning("未找到已发送文件夹，跳过保存")
            
            imap.logout()
        except Exception as imap_error:
            # IMAP保存失败不影响邮件发送，只记录警告
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"保存到已发送文件夹失败: {str(imap_error)}")
        
        return True
    except smtplib.SMTPException as e:
        raise ValueError(f"SMTP错误: {str(e)}")
    except Exception as e:
        raise ValueError(f"发送邮件失败: {str(e)}")
